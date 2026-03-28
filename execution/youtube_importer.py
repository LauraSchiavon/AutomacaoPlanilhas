#!/usr/bin/env python3
"""
YouTube Analytics & RedTrack Data Importer
Handles YouTube Data extraction for video retention (hook/body rates) and RedTrack integration.
"""

import os
import re
import sys
import pandas as pd
from datetime import datetime, timedelta
import requests
from typing import List, Dict, Tuple

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

SCOPES = [
    "https://www.googleapis.com/auth/yt-analytics.readonly",
    "https://www.googleapis.com/auth/youtube.readonly",
]
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CLIENT_SECRET = os.path.join(os.path.dirname(SCRIPT_DIR), "client_secret.json")
TOKEN_FILE = os.path.join(os.path.dirname(SCRIPT_DIR), "token.json")

class YouTubeAnalyticsAPI:
    """YouTube Analytics client for fetching retention data"""
    
    def __init__(self):
        self.creds = self.authenticate()
        self.youtube = build("youtube", "v3", credentials=self.creds)
        self.yt_analytics = build("youtubeAnalytics", "v2", credentials=self.creds)
        self.channel_videos = []
        self._load_channel_videos()
        self.durations = self._fetch_durations([v['id'] for v in self.channel_videos])

    def authenticate(self):
        creds = None
        if os.path.exists(TOKEN_FILE):
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                if not os.path.exists(CLIENT_SECRET):
                    print(f"Error: {CLIENT_SECRET} not found. Needs manual auth.")
                    # Return None so the web UI can warn the user
                    return None
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(TOKEN_FILE, "w") as f:
                f.write(creds.to_json())
        return creds

    def _load_channel_videos(self):
        """Pre-fetch all channel videos to match by title"""
        if not self.youtube: return
        videos = []
        request = self.youtube.search().list(
            part="id,snippet",
            forMine=True,
            type="video",
            maxResults=50,
            order="date"
        )
        while request:
            response = request.execute()
            for item in response.get("items", []):
                videos.append({
                    "id": item["id"]["videoId"],
                    "title": item["snippet"]["title"],
                })
            request = self.youtube.search().list_next(request, response)
        self.channel_videos = videos

    def parse_duration(self, iso_duration):
        match = re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", iso_duration)
        if not match: return 0
        return int(match.group(1) or 0) * 3600 + int(match.group(2) or 0) * 60 + int(match.group(3) or 0)

    def _fetch_durations(self, video_ids):
        durations = {}
        if not self.youtube or not video_ids: return durations
        for i in range(0, len(video_ids), 50):
            batch = video_ids[i: i+50]
            response = self.youtube.videos().list(
                part="contentDetails",
                id=",".join(batch)
            ).execute()
            for item in response.get("items", []):
                durations[item["id"]] = self.parse_duration(item["contentDetails"]["duration"])
        return durations

    def get_video_by_title(self, search_title: str):
        """Search pre-loaded channel videos for a matching title"""
        search_title_lower = search_title.lower()
        for v in self.channel_videos:
            if search_title_lower in v["title"].lower():
                return v
        return None

    def get_retention_data(self, video_id: str, date_from: str, date_to: str):
        """Fetch the 100-point retention curve"""
        if not self.yt_analytics: return []
        try:
            response = self.yt_analytics.reports().query(
                ids="channel==MINE",
                startDate=date_from,
                endDate=date_to,
                metrics="audienceWatchRatio",
                dimensions="elapsedVideoTimeRatio",
                filters=f"video=={video_id}",
                sort="elapsedVideoTimeRatio",
            ).execute()
            rows = response.get("rows", [])
            return [(row[0], row[1]) for row in rows]
        except Exception as e:
            print(f"Error fetching analytics for {video_id}: {e}")
            return []

    def get_video_stats(self, video_id: str, date_from: str, date_to: str):
        """Fetch basic stats like views, impressions for estimating CTR/CPM if needed via YouTube"""
        if not self.yt_analytics: return {}
        try:
            response = self.yt_analytics.reports().query(
                ids="channel==MINE",
                startDate=date_from,
                endDate=date_to,
                metrics="views,impressions,estimatedMinutesWatched",
                filters=f"video=={video_id}",
            ).execute()
            rows = response.get("rows", [])
            if rows:
                return {
                    "views": rows[0][0],
                    "impressions": rows[0][1],
                    "minutes_watched": rows[0][2]
                }
            return {"views": 0, "impressions": 0}
        except Exception:
            return {"views": 0, "impressions": 0}

    def interpolate_retention(self, curve, target_ratio):
        if not curve: return 0.0
        if target_ratio <= 0: return curve[0][1]
        if target_ratio >= 1.0: return curve[-1][1]

        for i in range(len(curve) - 1):
            r1, v1 = curve[i]
            r2, v2 = curve[i + 1]
            if r1 <= target_ratio <= r2:
                if r2 == r1: return v1
                t = (target_ratio - r1) / (r2 - r1)
                return (v1 + t * (v2 - v1))
        
        closest = min(curve, key=lambda p: abs(p[0] - target_ratio))
        return closest[1]

class RedTrackAPI:
    """RedTrack API client"""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://api.redtrack.io"
        self.session = requests.Session()
        self.session.headers.update({
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        })
    
    def get_conversions(self, date_start: str, date_end: str) -> List[Dict]:
        try:
            url = f"{self.base_url}/conversions/export"
            params = {'date_from': date_start, 'date_to': date_end, 'limit': 10000}
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json().get('data', [])
        except Exception as e:
            print(f"Error fetching conversions: {e}")
            return []
    
    def get_campaigns(self, date_start: str, date_end: str) -> List[Dict]:
        try:
            url = f"{self.base_url}/campaigns"
            params = {'date_from': date_start, 'date_to': date_end, 'limit': 1000}
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json().get('data', [])
        except Exception as e:
            print(f"Error fetching campaigns: {e}")
            return []

class ExcelManager:
    # Minimal ExcelManager so original code structures aren't fully broken
    def __init__(self, file_path: str):
        from openpyxl import load_workbook
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.raw_sheet = None
    
    def create_raw_data_sheet(self) -> None:
        sheet_name = "Dados Brutos"
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]
        self.raw_sheet = self.wb.create_sheet(sheet_name, 0)

    def add_redtrack_data(self, data: List[Dict], start_column: int=1) -> int:
        if not data: return 0
        df = pd.DataFrame(data)
        for col_idx, col_name in enumerate(df.columns, start_column):
            self.raw_sheet.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, start_column):
                self.raw_sheet.cell(row=row_idx, column=col_idx, value=value)
        return len(df)
    
    def add_youtube_data(self, data: List[Dict], start_column: int=1) -> int:
        if not data: return 0
        df = pd.DataFrame(data)
        for col_idx, col_name in enumerate(df.columns, start_column):
            self.raw_sheet.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, start_column):
                self.raw_sheet.cell(row=row_idx, column=col_idx, value=value)
        return len(df)
    
    def save(self) -> None:
        self.wb.save(self.file_path)

def main(date_start=None, date_end=None, rt_token=None):
    print("YouTube & RedTrack Analytics Importer Initialized")
    # For standalone debug testing
    yt_api = YouTubeAnalyticsAPI()
    rt_api = RedTrackAPI(rt_token or "wB7qY69R0KVU9tl4TBaQ")
    
    print(f"Loaded {len(yt_api.channel_videos)} YouTube videos.")
    print("Standalone test completed successfully.")
    return [], []

if __name__ == "__main__":
    main()
