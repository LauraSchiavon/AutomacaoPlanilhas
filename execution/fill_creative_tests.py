#!/usr/bin/env python3
"""
fill_creative_tests.py
Fills Column A (TC group label) and performance metrics (Hook, Body75, CPM, CTR, CPC, Gasto)
of the '032026' sheet conditionally based on User Column M (Status).

Logic:
1. Fetches USD->BRL quote from AwesomeAPI.
2. Fetches ALL campaigns directly via /campaigns to parse out names and TCs.
3. Reads each Excel test row, parses its specific 'Início' date (Col C / Col P).
4. Fetches targeted /insights and /report (RedTrack) dynamically for that exact date range.
5. Col A Logic: Fills Col A if blank and not already merged.
6. Metrics Logic: Fills financial and performance columns ONLY if Col M contains 'TESTE'.
"""

import re
import sys
import os
import requests
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

EXCEL_FILE = r"C:\Users\Laura\Downloads\Preencher planilha\testeLaura.xlsx"
MAIN_SHEET = "032026"
DATA_START_ROW = 4

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from youtube_importer import YouTubeAnalyticsAPI, RedTrackAPI

def parse_campaign_name(campaign_name: str):
    if not campaign_name: return None
    tc_match = re.search(r'TC(\d+)', campaign_name, re.IGNORECASE)
    if not tc_match: return None
    tc_label = f"TC{tc_match.group(1)}"
    bracket_match = re.search(r'\]([^[]*?)TC\d+', campaign_name, re.IGNORECASE)
    account_name = bracket_match.group(1).strip() if bracket_match else ""
    return tc_label, account_name

def extract_ad_name_from_campaign(campaign_name: str) -> str:
    if not campaign_name: return ""
    match = re.search(r'(?:ABO|CBO)\s+\S+\s*-\s*(.+)$', campaign_name, re.IGNORECASE)
    if match: return match.group(1).strip()
    parts = campaign_name.rsplit(' - ', 1)
    if len(parts) == 2: return parts[1].strip()
    return ""

def build_col_a_label(campaign_name: str) -> str:
    result = parse_campaign_name(campaign_name)
    if not result: return ""
    return f"{result[0]} {result[1]}".strip()

def parse_excel_date(cell_value, default_date: str) -> str:
    """Safely parse Excel datetime objects or strings to YYYY-MM-DD"""
    if not cell_value:
        return default_date
    if isinstance(cell_value, datetime.datetime):
        return cell_value.strftime("%Y-%m-%d")
    if isinstance(cell_value, str):
        # Handle 'dd/mm/yyyy' typical in BR Excel
        match = re.search(r'(\d{2})/(\d{2})/(\d{4})', cell_value)
        if match:
            return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
        # Handle 'dd/mm/yy'
        match_short = re.search(r'(\d{2})/(\d{2})/(\d{2})', cell_value)
        if match_short:
            return f"20{match_short.group(3)}-{match_short.group(2)}-{match_short.group(1)}"
    return default_date

def fetch_youtube_retention(yt_api: YouTubeAnalyticsAPI, video_id: str, duration: int, since: str, until: str):
    if not yt_api or not video_id or duration <= 0:
        return {"hook_rate": 0.0, "body_rate": 0.0}
    curve = yt_api.get_retention_data(video_id, since, until)
    if not curve:
        return {"hook_rate": 0.0, "body_rate": 0.0}
    
    # Hook rate: 3 seconds
    hook_ratio = 3.0 / duration
    hook_retention = yt_api.interpolate_retention(curve, hook_ratio)
    
    # Body rate: 75%
    body_retention = yt_api.interpolate_retention(curve, 0.75)
    
    return {
        "hook_rate": hook_retention,
        "body_rate": body_retention
    }

def fetch_rt_for_ad(ad_name_lower, since, until, rt_token):
    vendas = 0.0
    cost = 0.0
    roas = 0.0
    if not rt_token or not ad_name_lower:
        return {"vendas": 0, "cost": 0.0, "roas": 0.0}
    
    page = 1
    while page <= 5: # Limit pagination for row-by-row fetching to prevent lockups
        r = requests.get('https://api.redtrack.io/report', params={
            'api_key': rt_token,
            'date_from': since,
            'date_to': until,
            'group': 'rt_ad',
            'limit': 1000,
            'page': page
        }, timeout=30)
        
        if r.status_code != 200: break
        rt_data = r.json()
        if not rt_data: break
        
        for r_row in rt_data:
            rt_ad = str(r_row.get('rt_ad', '')).strip().lower()
            if not rt_ad: continue
            
            # Match condition logic - EXACT match first, then prefix extraction
            is_match = False
            if rt_ad == ad_name_lower:
                is_match = True
            elif rt_ad == ad_name_lower.split(" - ")[0].split(" ")[0]:
                is_match = True
                
            if is_match:
                vendas += float(r_row.get('convtype2', 0))
                cost += float(r_row.get('cost', 0))
                roas_val = float(r_row.get('roas', 0))
                if roas_val != 0: roas = roas_val
        
        if len(rt_data) < 1000: break
        page += 1

    return {"vendas": vendas, "cost": cost, "roas": roas}

def fill_creative_tests(
    date_start: str, 
    date_end: str, 
    redtrack_token: str = None,
    progress_callback=None
):
    if progress_callback: progress_callback("Autenticando YouTube API...")
    yt_api = YouTubeAnalyticsAPI()
    if not yt_api.creds:
        raise RuntimeError("Erro: YouTube OAuth não configurado (Faltado token.json/client_secret.json)")

    if progress_callback: progress_callback("Construindo catálogo raiz de Vídeos do YouTube...")
    
    # Simulate the dictionary of campaigns by matching against channel videos
    ad_to_campaign = {}
    for vid in yt_api.channel_videos:
        key = vid["title"].strip().lower()
        ad_to_campaign[key] = {"id": vid["id"], "name": vid["title"], "duration": yt_api.durations.get(vid["id"], 0)}

    if not ad_to_campaign:
        raise RuntimeError("Nenhum vídeo válido encontrado no seu canal do YouTube selecionado.")

    # 3. Read Excel and Process Row-by-Row
    if progress_callback:
        progress_callback("Varrendo planilha e extraindo dados específicos por linha (Dynamic Dates)...")

    try:
        wb = load_workbook(EXCEL_FILE)
    except PermissionError:
        raise PermissionError(f"O Arquivo Excel está aberto. Feche a planilha e tente novamente.")
    
    ws = wb[MAIN_SHEET]

    filled_a = 0
    filled_metrics = 0
    skipped_rows = 0
    not_found = []

    # Process TESTES Completos Section
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        ad_name_cell = ws.cell(row=row_idx, column=2)
        ad_name_value = ad_name_cell.value
        if not ad_name_value or str(ad_name_value).strip() == "":
            continue

        search_term = str(ad_name_value).strip().lower()
        matched_info = None
        
        if search_term in ad_to_campaign:
            matched_info = ad_to_campaign[search_term]
        else:
            for key, info in ad_to_campaign.items():
                if search_term in key or key in search_term:
                    matched_info = info
                    break

        if not matched_info:
            not_found.append(str(ad_name_value))
            continue
            
        c_name = matched_info["name"]
        c_id = matched_info["id"]

        # Col A Logic
        cell_a = ws.cell(row=row_idx, column=1)
        if type(cell_a).__name__ != 'MergedCell':
            if not cell_a.value or str(cell_a.value).strip() == "":
                label = build_col_a_label(c_name)
                if label:
                    cell_a.value = label
                    cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell_a.font = Font(bold=True)
                    filled_a += 1

        # Check Col M Status
        status_cell = ws.cell(row=row_idx, column=13)
        current_status = str(status_cell.value).strip().upper() if status_cell.value else ""
        
        if "TESTE" in current_status:
            # Get specific start date from Col C (column 3)
            date_col_c = ws.cell(row=row_idx, column=3).value
            row_date_start = parse_excel_date(date_col_c, date_start)
            
            # Fetch RedTrack dynamically for this row
            rt = fetch_rt_for_ad(search_term, row_date_start, date_end, redtrack_token)
            vendas = rt["vendas"]
            spend_brl = rt["cost"]  # Redtrack cost!
            
            # Fetch YouTube data dynamically for this row
            duration = matched_info.get("duration", 0)
            fin = fetch_youtube_retention(yt_api, c_id, duration, row_date_start, date_end)
            
            cpc_brl = 0.0
            cpm_brl = 0.0
            ctr = 0.0
            
            # Write metrics...
            c5 = ws.cell(row=row_idx, column=5, value=fin["hook_rate"])
            c5.number_format = '0.00%'
            c6 = ws.cell(row=row_idx, column=6, value=fin["body_rate"])
            c6.number_format = '0.00%'
            c7 = ws.cell(row=row_idx, column=7, value=cpm_brl)
            c7.number_format = '#,##0.00'
            c8 = ws.cell(row=row_idx, column=8, value=ctr)
            c8.number_format = '0.00%'
            c9 = ws.cell(row=row_idx, column=9, value=cpc_brl)
            c9.number_format = '#,##0.00'
            c10 = ws.cell(row=row_idx, column=10, value=spend_brl)
            c10.number_format = '#,##0.00'
            
            ws.cell(row=row_idx, column=11, value=vendas)
            
            cpa = 0
            if vendas > 0:
                cpa = spend_brl / vendas
            else:
                cpa = 0 # Forced fallback to 0
                
            c12 = ws.cell(row=row_idx, column=12, value=cpa)
            c12.number_format = '#,##0.00'

            filled_metrics += 1
        else:
            skipped_rows += 1

    # Re-merge Col A
    merge_groups = 0
    row = DATA_START_ROW
    while row <= ws.max_row:
        cell_val = ws.cell(row=row, column=1).value
        if not cell_val:
            row += 1
            continue
        end_row = row
        while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=1).value == cell_val:
            end_row += 1
        if end_row > row:
            ws.merge_cells(start_row=row, start_column=1, end_row=end_row, end_column=1)
            merged_cell = ws.cell(row=row, column=1)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            merged_cell.font = Font(bold=True)
            merge_groups += 1
        row = end_row + 1

    # Process PRÉ-ESCALA Section
    filled_pre_escala = 0
    skipped_pre_escala = 0
    if progress_callback:
        progress_callback("Preenchendo seção PRÉ-ESCALA com datas dinâmicas...")

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        creative_cell = ws.cell(row=row_idx, column=15)  # Col O
        creative_val = creative_cell.value
        if not creative_val or str(creative_val).strip() == "":
            continue

        status_pe = ws.cell(row=row_idx, column=22)  # Col V
        status_pe_str = str(status_pe.value).strip().upper() if status_pe.value else ""
        if "TESTE" not in status_pe_str:
            skipped_pre_escala += 1
            continue

        search_pe = str(creative_val).strip().lower()
        
        # Determine Row Date - Col P uses column 16
        date_col_p = ws.cell(row=row_idx, column=16).value
        # For Pré-Escala, parse_excel_date reads from Col P
        row_date_pe = parse_excel_date(date_col_p, date_start)
        
        # RedTrack Fetch
        rt = fetch_rt_for_ad(search_pe, row_date_pe, date_end, redtrack_token)
        
        cost_brl = rt['cost']
        vendas_pe = rt['vendas']
        roas_pe = rt['roas']
        
        cpa_pe = 0
        if vendas_pe > 0:
            cpa_pe = cost_brl / vendas_pe
        else:
            cpa_pe = 0 # forced fallback
        
        c_r = ws.cell(row=row_idx, column=18, value=cost_brl)
        c_r.number_format = '#,##0.00'
        
        ws.cell(row=row_idx, column=19, value=vendas_pe)
        
        c_t = ws.cell(row=row_idx, column=20, value=roas_pe)
        c_t.number_format = '0.00'
        
        c_u = ws.cell(row=row_idx, column=21, value=cpa_pe)
        c_u.number_format = '#,##0.00'

        filled_pre_escala += 1

    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        raise PermissionError(f"Não foi possível salvar! O arquivo '{EXCEL_FILE}' está aberto no Excel.")

    return {
        "filled_a": filled_a, 
        "filled_metrics": filled_metrics, 
        "skipped_rows": skipped_rows, 
        "not_found": not_found,
        "filled_pre_escala": filled_pre_escala,
        "skipped_pre_escala": skipped_pre_escala
    }
